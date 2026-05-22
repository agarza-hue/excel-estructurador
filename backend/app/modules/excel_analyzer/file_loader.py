"""
Universal Excel/spreadsheet loader.
Normalizes any supported format to a temporary .xlsx file
so the openpyxl-based analyzer can process it unchanged.

Supported formats:
  .xlsx  .xlsm  .xltx  .xltm  — native openpyxl (pass-through)
  .xls                         — xlrd -> openpyxl conversion
  .xlsb                        — pyxlsb -> openpyxl conversion
  .ods                         — odfpy via pandas -> openpyxl conversion
  .csv   .tsv                  — pandas -> synthetic single-sheet workbook
"""

import tempfile
import shutil
from pathlib import Path
import openpyxl
import pandas as pd
import structlog

log = structlog.get_logger()

# Formats that openpyxl handles natively
NATIVE = {".xlsx", ".xlsm", ".xltx", ".xltm"}
ALL_SUPPORTED = NATIVE | {".xls", ".xlsb", ".ods", ".csv", ".tsv"}


def normalize_to_xlsx(src_path: str) -> tuple[str, bool]:
    """
    Returns (path_to_use, was_converted).
    If was_converted=True, caller must delete the temp file when done.
    """
    p = Path(src_path)
    ext = p.suffix.lower()

    if ext in NATIVE:
        return src_path, False

    if ext not in ALL_SUPPORTED:
        raise ValueError(f"Unsupported file format: '{ext}'. Supported: {sorted(ALL_SUPPORTED)}")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    out_path = tmp.name

    try:
        if ext == ".xls":
            _xls_to_xlsx(src_path, out_path)
        elif ext == ".xlsb":
            _xlsb_to_xlsx(src_path, out_path)
        elif ext == ".ods":
            _ods_to_xlsx(src_path, out_path)
        elif ext in (".csv", ".tsv"):
            _csv_to_xlsx(src_path, out_path, ext)
        else:
            shutil.copy2(src_path, out_path)

        log.info("file_normalized", src=p.name, ext=ext, out=out_path)
        return out_path, True

    except Exception as e:
        Path(out_path).unlink(missing_ok=True)
        raise RuntimeError(f"Cannot convert {p.name} ({ext}): {e}") from e


# ── Converters ────────────────────────────────────────────────────────────────

def _xls_to_xlsx(src: str, dst: str):
    import xlrd
    rb = xlrd.open_workbook(src, formatting_info=False)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in rb.sheet_names():
        rs = rb.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                cell = rs.cell(row_idx, col_idx)
                value = _xlrd_value(cell, rb)
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    wb.save(dst)


def _xlrd_value(cell, workbook):
    import xlrd
    t = cell.ctype
    if t == xlrd.XL_CELL_EMPTY:
        return None
    if t == xlrd.XL_CELL_TEXT:
        return cell.value
    if t == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return int(v) if v == int(v) else v
    if t == xlrd.XL_CELL_DATE:
        try:
            from xlrd import xldate_as_datetime
            return xldate_as_datetime(cell.value, workbook.datemode)
        except Exception:
            return cell.value
    if t == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if t == xlrd.XL_CELL_ERROR:
        return None
    return cell.value


def _xlsb_to_xlsx(src: str, dst: str):
    from pyxlsb import open_workbook as open_xlsb
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    with open_xlsb(src) as wb:
        for name in wb.sheets:
            ws_out = wb_out.create_sheet(title=name)
            with wb.get_sheet(name) as ws:
                for row in ws.rows():
                    for cell in row:
                        if cell.v is not None:
                            ws_out.cell(row=cell.r + 1, column=cell.c + 1, value=cell.v)

    wb_out.save(dst)


def _ods_to_xlsx(src: str, dst: str):
    # pandas reads .ods via odfpy
    try:
        sheets = pd.read_excel(src, sheet_name=None, engine="odf")
    except Exception:
        # fallback: treat as single sheet
        df = pd.read_excel(src, engine="odf")
        sheets = {"Sheet1": df}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _dfs_to_wb(sheets, wb, dst)


def _csv_to_xlsx(src: str, dst: str, ext: str):
    sep = "\t" if ext == ".tsv" else ","
    try:
        df = pd.read_csv(src, sep=sep, dtype=str, encoding="utf-8")
    except UnicodeDecodeError:
        df = pd.read_csv(src, sep=sep, dtype=str, encoding="latin-1")

    sheet_name = Path(src).stem[:31] or "Sheet1"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(val) else val)

    wb.save(dst)


def _dfs_to_wb(sheets: dict, wb: openpyxl.Workbook, dst: str):
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(val) else val)
    wb.save(dst)

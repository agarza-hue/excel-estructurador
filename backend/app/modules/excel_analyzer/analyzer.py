import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Any
import structlog

from .type_inferrer import infer_column_type, suggest_sql_type
from .formula_parser import extract_formula_map, parse_formula
from .structure_detector import (
    detect_header_row, detect_table_regions, classify_sheet_type,
    detect_relationships, sanitize_column_name
)

log = structlog.get_logger()
SAMPLE_ROWS = 20
MAX_COL_SCAN = 200


class ExcelAnalyzer:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb = None
        self._tmp_path = None  # path to converted temp file (if any)

    def analyze(self, on_progress=None, collect_decisions: bool = True) -> dict:
        """
        on_progress(pct: int, msg: str, level: str) — optional callback
        called after each sheet is processed. pct is in range 15–55.
        """
        from .file_loader import normalize_to_xlsx
        work_path, was_converted = normalize_to_xlsx(str(self.file_path))
        if was_converted:
            self._tmp_path = work_path

        log.info("analyzing_workbook", path=self.file_path.name, format=self.file_path.suffix.lower())
        self.wb = openpyxl.load_workbook(
            work_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )

        from .decision_collector import DecisionCollector
        collector = DecisionCollector() if collect_decisions else None

        sheets_meta = []
        formula_map = {}
        total_rows = 0
        total_cols = 0
        total_formulas = 0
        hidden_count = 0
        pivot_count = 0
        named_ranges = self._extract_named_ranges()
        total_sheets = len(self.wb.sheetnames)

        for idx, sheet_name in enumerate(self.wb.sheetnames):
            ws = self.wb[sheet_name]
            log.info("analyzing_sheet", sheet=sheet_name, index=idx)

            is_hidden = ws.sheet_state != "visible"
            if is_hidden:
                hidden_count += 1

            raw_data = self._read_sheet_data(ws)
            if not raw_data:
                sheets_meta.append(self._empty_sheet_meta(idx, sheet_name, is_hidden))
                continue

            header_row_idx = detect_header_row(raw_data)
            headers = self._extract_headers(raw_data, header_row_idx)
            data_rows = raw_data[header_row_idx + 1:]

            col_meta = self._analyze_columns(headers, data_rows)
            formula_data = extract_formula_map(raw_data, sheet_name)
            merged = self._extract_merged_ranges(ws)
            has_pivot = self._detect_pivot(ws)

            # Use actual header count as col_count (not ws.max_column which includes empty cols)
            row_count = max(len(raw_data), 0)
            col_count = len(col_meta) if col_meta else (ws.max_column or 0)
            total_rows += row_count
            total_cols = max(total_cols, col_count)
            total_formulas += formula_data["formula_count"]
            if has_pivot:
                pivot_count += 1

            sheet_meta = {
                "index": idx,
                "sheet_name": sheet_name,
                "is_hidden": is_hidden,
                "row_count": row_count,
                "col_count": col_count,
                "header_row": header_row_idx,
                "data_start_row": header_row_idx + 1,
                "has_merged_cells": len(merged) > 0,
                "has_formulas": formula_data["formula_count"] > 0,
                "has_pivot_table": has_pivot,
                "formula_count": formula_data["formula_count"],
                "columns_meta": col_meta,
                "merged_ranges": merged,
                "formula_summary": {
                    "categories": formula_data["categories"],
                    "cross_sheet_deps": formula_data["cross_sheet_dependencies"],
                    "has_financial": formula_data["has_financial"],
                    "has_lookup": formula_data["has_lookup"],
                    "complexity": formula_data["complexity_distribution"],
                },
                "sample_data": self._build_sample(headers, data_rows),
                "regions": detect_table_regions(raw_data),
            }

            sheet_type, confidence = classify_sheet_type(sheet_meta)
            sheet_meta["sheet_type"] = sheet_type
            sheet_meta["confidence_score"] = confidence

            sheets_meta.append(sheet_meta)
            formula_map[sheet_name] = formula_data

            if collector:
                collector.collect_sheet(sheet_meta)

            if on_progress:
                pct = 15 + int(((idx + 1) / total_sheets) * 40)  # 15 → 55 %
                on_progress(pct, f"Hoja {idx + 1}/{total_sheets}: {sheet_name} "
                            f"({row_count:,} filas, {col_count} cols)")

        if collector:
            collector.collect_workbook(sheets_meta)

        relationships = detect_relationships(sheets_meta)

        workbook_structure = {
            "filename": self.file_path.name,
            "sheet_count": len(self.wb.sheetnames),
            "sheet_names": self.wb.sheetnames,
            "total_rows": total_rows,
            "total_columns": total_cols,
            "has_formulas": total_formulas > 0,
            "formula_count": total_formulas,
            "has_pivot_tables": pivot_count > 0,
            "pivot_table_count": pivot_count,
            "has_named_ranges": len(named_ranges) > 0,
            "named_ranges": named_ranges,
            "has_hidden_sheets": hidden_count > 0,
            "hidden_sheet_count": hidden_count,
            "sheets": sheets_meta,
            "relationships": relationships,
        }

        inferred_entities = self._infer_entities(sheets_meta)

        # Clean up temp converted file
        if self._tmp_path:
            try:
                Path(self._tmp_path).unlink(missing_ok=True)
            except Exception:
                pass

        return {
            "workbook_structure": workbook_structure,
            "formula_map": formula_map,
            "inferred_entities": inferred_entities,
            "decisions": collector.to_list() if collector else [],
            "summary": {
                "sheet_count": len(sheets_meta),
                "total_rows": total_rows,
                "total_formulas": total_formulas,
                "entity_count": len(inferred_entities),
                "relationship_count": len(relationships),
                "decision_count": collector.count() if collector else 0,
            },
        }

    def _read_sheet_data(self, ws) -> list[list[Any]]:
        rows = []
        for row in ws.iter_rows(max_col=MAX_COL_SCAN, values_only=False):
            row_data = []
            for cell in row:
                # Capture formula string, not computed value
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    row_data.append(cell.value)  # formula string
                else:
                    row_data.append(cell.value)
            rows.append(row_data)
        # Trim trailing empty rows
        while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
            rows.pop()
        return rows

    def _extract_headers(self, rows: list, header_idx: int) -> list[str]:
        if header_idx >= len(rows):
            return []
        raw = rows[header_idx]

        # Find last non-empty column in header row
        last_col = 0
        for i, cell in enumerate(raw):
            if cell is not None and str(cell).strip():
                last_col = i
        raw = raw[:last_col + 1]

        seen: dict[str, int] = {}
        headers = []
        for cell in raw:
            name = sanitize_column_name(str(cell) if cell is not None else "")
            if not name:
                name = "col_unknown"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            headers.append(name)
        return headers

    def _analyze_columns(self, headers: list[str], data_rows: list[list]) -> list[dict]:
        col_meta = []
        if not data_rows or not headers:
            return col_meta

        n_cols = len(headers)
        for col_idx in range(n_cols):
            values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
            meta = infer_column_type(values)
            meta["name"] = headers[col_idx]
            meta["index"] = col_idx
            meta["sql_type"] = suggest_sql_type(meta)
            meta["sample_values"] = [str(v) for v in values[:5] if v is not None]
            col_meta.append(meta)

        return col_meta

    def _extract_merged_ranges(self, ws) -> list[str]:
        return [str(r) for r in ws.merged_cells.ranges]

    def _detect_pivot(self, ws) -> bool:
        try:
            return len(ws._pivots) > 0 if hasattr(ws, "_pivots") else False
        except Exception:
            return False

    def _extract_named_ranges(self) -> list[dict]:
        ranges = []
        try:
            for name, defn in self.wb.defined_names.items():
                destinations = list(defn.destinations)
                ranges.append({
                    "name": name,
                    "destinations": [{"sheet": s, "range": c} for s, c in destinations],
                })
        except Exception:
            pass
        return ranges

    def _build_sample(self, headers: list, data_rows: list) -> list[dict]:
        sample = []
        for row in data_rows[:SAMPLE_ROWS]:
            record = {}
            for i, h in enumerate(headers):
                record[h] = str(row[i]) if i < len(row) and row[i] is not None else None
            sample.append(record)
        return sample

    def _empty_sheet_meta(self, idx: int, name: str, hidden: bool) -> dict:
        return {
            "index": idx,
            "sheet_name": name,
            "is_hidden": hidden,
            "row_count": 0,
            "col_count": 0,
            "sheet_type": "empty",
            "confidence_score": 1.0,
            "columns_meta": [],
            "formula_count": 0,
        }

    def _infer_entities(self, sheets_meta: list[dict]) -> list[dict]:
        entities = []
        for sheet in sheets_meta:
            if sheet.get("sheet_type") in ("transactional", "master"):
                cols = sheet.get("columns_meta", [])
                pk_candidates = [c for c in cols if c.get("is_unique") and
                                 c.get("null_rate", 1) == 0]
                fk_candidates = [c for c in cols if c["name"].endswith("_id") or
                                 c["name"].endswith("_codigo") or c["name"].endswith("_code")]

                entities.append({
                    "entity_name": sheet["sheet_name"],
                    "source_sheet": sheet["sheet_name"],
                    "entity_type": sheet["sheet_type"],
                    "column_count": len(cols),
                    "row_estimate": sheet.get("row_count", 0),
                    "pk_candidates": [c["name"] for c in pk_candidates],
                    "fk_candidates": [c["name"] for c in fk_candidates],
                    "columns": cols,
                })
        return entities
